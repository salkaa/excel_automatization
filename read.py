import openpyxl
import pandas as pd

def read_data():

    # workbook = openpyxl.load_workbook('/home/salka/Documents/excel_files/new_plan.xlsx')
    # sheet = workbook.active
    df_sheet_name = pd.read_excel('/home/salka/Documents/excel_files/new_plan.xlsx', sheet_name=1)
    print(df_sheet_name)


    # max_col = sheet.max_column
    # max_row = sheet.max_row

    # chcem vypocitat kde su vsetky tabulky a dat ich do pandas


    # data = [] # mam 2d list so vsetkymi bunkami, ktore su string

    # for i in range(2, max_row+1):
    #     value_row = []
    #     for n in range(1, max_col+1):
    #         x = sheet.cell(row = i, column = n)
    #         if isinstance(x.value, str):
    #             value_row.append(x.value)
    #     data.append(value_row)




    # headers_of_weeks = [] # mam 2d list s nazvami tyzdnov a ich indexami v data
    # indexes = []
    # for value in data:
    #     if len(value) == 1 and "Týždeň" in value[0]:
    #         indexes.append(data.index(value))
    #         headers_of_weeks.append([value[0], data.index(value)])
    




    # weeks_dict = {} # data rozdelene podla indexov headers_of_weeks (bez nich)

    # for index in indexes[:-1]:
    #     name = headers_of_weeks[indexes.index(index)][0]
    #     weeks_dict[name] = [data[i] for i in range(index+1, indexes[indexes.index(index)+1])]





    # # key je nazov treningu, values su nazvy cvikov, pocet serii opakovani a hmotnost, komentar

    # days = {}

    # for key in weeks_dict:
        
    #     values = weeks_dict[key]
    #     names = []
    #     for element in values[0]:
        
    #         date_name = element.split() # date_name = ['datum', 'meno treningu']
    #         name = date_name[1] # 'meno treningu'
    #         names.append(name)
    #         if date_name[1] not in days.keys():
    #             days[name] = []
        
    #     divided = []
    #     for x in values[0]:
    #         divided.append([])
    #     reps = []

    #     for array in values[1:]:
            
    #         if len(array) == 3:
    #             reps = array
            
    #         else:
    #             for i in range(len(array)):
    #                 if i%3 == 1:
    #                     divided[i//3].append(reps[i//3] + array[i])
    #                 else:
    #                     divided[i//3].append(array[i])
        
        
    #     for name in names:
    #         days[name].append(divided[names.index(name)])




    # # potrebujem dictionary vsetkych cvikov ako keys a values budu opakovania a komentare v 2d lister

    # exercises = {}

    # for key in days:
    #     n = days[key]
    #     for array in n:
    #         for i in range(0, len(array), 3):
    #             if array[i] not in exercises:
    #                 exercises[array[i]] = [[array[i+1], array[i+2]]]
    #             else:
    #                 exercises[array[i]].append([array[i+1], array[i+2]])

    # for key in exercises:
    #     print(key, exercises[key])
          
        


    # return data, headers_of_weeks, weeks_dict

read_data()

